"""
Master Price Analysis Tool
--------------------------
Usage: python master_price_tool.py <offer_file.xlsx> [output_name.xlsx]

Compares an offer/inventory file (UPC + price columns) against all configured
wholesaler pricelists and produces a formatted Excel report.

WHOLESALER CONFIG — add new wholesalers here:
  Each entry: (name, file_path, upc_col, price_col, header_row, sheet)
  - header_row: 0-based row index for pandas header= param
  - upc_col / price_col: 0-based column index
"""

import sys, os, re, glob, pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ════════════════════════════════════════════════════════════════
# WHOLESALER REGISTRY — edit this to add/remove wholesalers
# Format: (display_name, file_path_or_glob, upc_col, price_col, header_row, sheet_name_or_None)
# For multiple files of same wholesaler (different dates), use glob pattern — most recent price wins
# ════════════════════════════════════════════════════════════════
UPLOAD_DIR = '/mnt/user-data/uploads'
CLAUDE_DIR = '/home/claude'

WHOLESALERS = [
    {
        'name': 'MTZ',
        'files': sorted(glob.glob(f'{UPLOAD_DIR}/MTZpricelist__*.xlsx')),  # all files, most recent price wins
        'upc_col': 3,
        'price_col': 4,
        'header_row': 4,
        'sheet': 'Price',
    },
    {
        'name': 'Nandansons',
        'files': sorted(glob.glob(f'{UPLOAD_DIR}/Nandansons_Wholesale_Price_List030426.xlsx')) +
                 sorted(glob.glob(f'{CLAUDE_DIR}/Nandansons_Wholesale_Price_List__*.xlsx')) +
                 sorted(glob.glob(f'{CLAUDE_DIR}/Nandansons_Wholesale_Price_List081325*.xlsx')),
        'upc_col': 0,
        'price_col': 'auto',
        'header_row': 0,
        'sheet': None,
        'date_from_filename': True,
    },
    {
        'name': 'PCA',
        'files': [
            (f'{CLAUDE_DIR}/PRICE-LIST-10-08-2025__1_.xlsx', '2025-10-08'),
        ],
        'upc_col': 6,
        'price_col': 8,
    },
    {
        'name': 'GE',
        'files': [
            # (filepath, date) — all same structure: header row 0, UPC col 1, Price col 4
            (f'{CLAUDE_DIR}/GE_WHOLESALE_05_11.xlsx',       '2024-05-11'),
            (f'{CLAUDE_DIR}/GE_WHOLESALE_06-02.xlsx',       '2024-06-02'),
            (f'{CLAUDE_DIR}/GE_WHOLESALE_08-17_-.xlsx',     '2024-08-17'),
            (f'{CLAUDE_DIR}/GE_WHOLESALE_08-17_-__1_.xlsx', '2024-08-17'),
            (f'{CLAUDE_DIR}/WHOLESALE_08-31-.xlsx',         '2024-08-31'),
            (f'{CLAUDE_DIR}/WHOLESALE_12_21.xlsx',          '2024-12-21'),
            (f'{CLAUDE_DIR}/WHOLESALE_12_21__1_.xlsx',      '2024-12-21'),
            (f'{CLAUDE_DIR}/WHOLESALE_12-28.xlsx',          '2024-12-28'),
            (f'{CLAUDE_DIR}/GE_WHOLESALE_01-18.xlsx',       '2025-01-18'),
            (f'{CLAUDE_DIR}/GE_WHOLESALE_02-15_.xlsx',      '2025-02-15'),
            (f'{CLAUDE_DIR}/GE__WHOLESALE_03_01__1_.xlsx',  '2025-03-01'),
            (f'{CLAUDE_DIR}/Ge_Wholesale_03_16.xlsx',       '2025-03-16'),
            (f'{CLAUDE_DIR}/GE_wholesale_03__30.xlsx',      '2025-03-30'),
            (f'{CLAUDE_DIR}/GE_WHOLESALE_04_20.xlsx',       '2025-04-20'),
            (f'{CLAUDE_DIR}/GE_Wholesale_04-27.xlsx',       '2026-04-27'),
        ],
        'upc_col': 1,
        'price_col': 4,
    },
    {
        'name': 'PTC',
        'files': [
            # (filepath, sheet_name, header_row, date)
            (f'{UPLOAD_DIR}/PTC_NEW_ARRIVALS_03_04_26.xlsx',                                                     'Price List', 11, '2026-03-04'),
            (f'{UPLOAD_DIR}/PTC_NEW_ARRIVALS_02_18_26.xlsx',                                                     'Price List', 11, '2026-02-18'),
            (f'{UPLOAD_DIR}/PTC_NEW_ARRIVALS_12_08_25.xlsx',                                                     'Price List', 11, '2025-12-08'),
            (f'{UPLOAD_DIR}/PTC_SPECIAL_PRICE_LIST_AND_NEW_ARRIVALS_06_16_25.xlsx',                              'Sheet1',      8, '2025-06-16'),
            (f'{UPLOAD_DIR}/PTC_SPECIAL_PRICE_LIST_AND_NEW_ARRIVALS_05_29_25.xlsx',                              'Sheet1',      8, '2025-05-29'),
            (f'{UPLOAD_DIR}/PTC_SPECIAL_PRICE_LIST_AND_NEW_ARRIVALS_05_23_25.xlsx',                              'Sheet1',      8, '2025-05-23'),
            (f'{UPLOAD_DIR}/PTC_SPECIAL_PRICE_LIST_AND_NEW_ARRIVALS_04_24_25.xlsx',                              'Sheet1',      8, '2025-04-24'),
            (f'{UPLOAD_DIR}/PTC_SPECIAL_PRICE_LIST_AND_NEW_ARRIVALS_03_20_25.xlsx',                              'Sheet1',      8, '2025-03-20'),
            (f'{UPLOAD_DIR}/PTC_OFFER_01_31_25_NEW_ARRIVALS_AND_PRICE_LIST_SPECIALS__1_-2__1_.xlsx',             'Sheet1',      7, '2025-01-31'),
            (f'{UPLOAD_DIR}/PTC_NEW_ARRIVALS_AND_PRICE_LIST_10-22-24.xlsx',                                      'Sheet2',      6, '2024-10-22'),
        ],
        'upc_col': 1,   # UPC/EAN is column index 1
        'price_col': 4, # PRICE is column index 4
    },
]

# ════════════════════════════════════════════════════════════════
# OFFER FILE CONFIG
# Detect UPC + price columns automatically, or override here
# ════════════════════════════════════════════════════════════════
OFFER_UPC_COL   = 'auto'  # column name or 0-based index; 'auto' = first col with 'UPC/EAN/item'
OFFER_PRICE_COL = 'auto'  # column name or 0-based index; 'auto' = first col with 'price'
OFFER_DESC_COL  = 'auto'  # 'auto' = first col with 'desc/name/product'
OFFER_QTY_COL   = 'auto'  # 'auto' = first col with 'qty/quantity'


def load_mtz(cfg):
    """Load all MTZ pricelists → {upc: price} using most recent price per UPC"""
    all_rows = []
    for f in cfg['files']:
        if not os.path.exists(f):
            continue
        # Extract date from row 2, col 2
        raw = pd.read_excel(f, sheet_name=cfg['sheet'], header=None)
        date = str(raw.iloc[2, 2])[:10]

        df = pd.read_excel(f, sheet_name=cfg['sheet'], header=cfg['header_row'])
        df.columns = [str(c) for c in df.columns]
        upc_col   = df.columns[cfg['upc_col']]
        price_col = df.columns[cfg['price_col']]
        df = df.dropna(subset=[upc_col])
        df[upc_col]   = df[upc_col].astype(str).str.strip()
        df[price_col] = pd.to_numeric(df[price_col], errors='coerce')
        tmp = pd.DataFrame({'UPC': df[upc_col], 'Price': df[price_col], 'Date': date})
        all_rows.append(tmp.dropna(subset=['Price']))

    if not all_rows:
        return {}
    combined = pd.concat(all_rows, ignore_index=True).sort_values('Date', ascending=False)
    deduped  = combined.drop_duplicates(subset='UPC', keep='first')
    return dict(zip(deduped['UPC'], deduped['Price']))


def get_nan_date(filename):
    """Extract sortable date from Nandansons filename"""
    import re
    # 030426 → 2026-03-04
    m = re.search(r'(\d{6})(?:_|\.)', os.path.basename(filename))
    if m:
        d = m.group(1)
        return f"20{d[4:6]}-{d[0:2]}-{d[2:4]}"
    # __1_ style — map by number
    m2 = re.search(r'__(\d+)_', os.path.basename(filename))
    if m2:
        num = int(m2.group(1))
        # Approximate dates based on what we know
        dates = {1:'2026-02-07',2:'2026-02-01',3:'2026-01-10',4:'2025-12-13',
                 5:'2025-10-19',6:'2025-09-25',7:'2025-09-13',8:'2025-08-22'}
        return dates.get(num, '2025-01-01')
    return '2025-01-01'


def load_nandansons(cfg):
    """Load all Nandansons files, keep most recent price per UPC → {upc: price}"""
    rows = []
    for f in cfg['files']:
        if not os.path.exists(f): continue
        date = get_nan_date(f)
        df = pd.read_excel(f, header=cfg['header_row'])
        upc = df.iloc[:,cfg['upc_col']].astype(str).str.strip()
        # Detect price col: newer format has 8 cols (price at 7), older has 7 (price at 5)
        ncols = df.shape[1]
        pcol = 7 if ncols >= 8 else 5
        price = pd.to_numeric(df.iloc[:,pcol], errors='coerce')
        tmp = pd.DataFrame({'UPC': upc, 'Price': price, 'Date': date})
        rows.append(tmp.dropna(subset=['Price']))
    if not rows:
        return {}
    combined = pd.concat(rows).sort_values('Date', ascending=False)
    deduped = combined.drop_duplicates(subset='UPC', keep='first')
    return dict(zip(deduped['UPC'], deduped['Price']))


def load_ptc(cfg):
    """Load all PTC files (each has its own sheet/header), keep most recent price per UPC"""
    rows = []
    upc_col   = cfg['upc_col']
    price_col = cfg['price_col']
    for (fpath, sheet, hrow, date) in cfg['files']:
        if not os.path.exists(fpath):
            continue
        df = pd.read_excel(fpath, sheet_name=sheet, header=hrow)
        upc   = df.iloc[:, upc_col].apply(
            lambda v: str(int(float(v))) if pd.notna(v) and str(v).replace('.','').replace('e+','').replace('E+','').isdigit()
            else str(v)
        ).str.strip()
        price = pd.to_numeric(df.iloc[:, price_col], errors='coerce')
        desc  = df.iloc[:, 0].astype(str).str.strip()
        tmp   = pd.DataFrame({'UPC': upc, 'Price': price, 'Desc': desc, 'Date': date})
        tmp   = tmp.dropna(subset=['Price'])
        tmp   = tmp[tmp['UPC'].str.len() >= 8]  # filter junk
        rows.append(tmp)
    if not rows:
        return {}
    combined = pd.concat(rows, ignore_index=True).sort_values('Date', ascending=False)
    deduped  = combined.drop_duplicates(subset='UPC', keep='first')
    return dict(zip(deduped['UPC'], deduped['Price']))


def load_ge(cfg):
    """Load all GE wholesale files → {upc: price} using most recent price per UPC"""
    rows = []
    upc_col   = cfg['upc_col']
    price_col = cfg['price_col']
    for (fpath, date) in cfg['files']:
        if not os.path.exists(fpath):
            continue
        df = pd.read_excel(fpath, header=0)
        upc = df.iloc[:, upc_col].apply(
            lambda v: str(int(float(v))) if pd.notna(v) and str(v).replace('.','').replace('e+','').replace('E+','').isdigit()
            else str(v)
        ).str.strip()
        price = pd.to_numeric(df.iloc[:, price_col], errors='coerce')
        tmp = pd.DataFrame({'UPC': upc, 'Price': price, 'Date': date})
        tmp = tmp.dropna(subset=['Price'])
        tmp = tmp[tmp['UPC'].str.len() >= 8]
        rows.append(tmp)
    if not rows:
        return {}
    combined = pd.concat(rows, ignore_index=True).sort_values('Date', ascending=False)
    deduped  = combined.drop_duplicates(subset='UPC', keep='first')
    return dict(zip(deduped['UPC'], deduped['Price']))


def load_pca(cfg):
    """Load all PCA price list files → {upc: price} using most recent price per UPC"""
    rows = []
    upc_col   = cfg['upc_col']
    price_col = cfg['price_col']
    for (fpath, date) in cfg['files']:
        if not os.path.exists(fpath):
            continue
        df = pd.read_excel(fpath, header=0)
        upc = df.iloc[:, upc_col].apply(
            lambda v: str(int(float(v))) if pd.notna(v) and str(v).replace('.','').replace('e+','').replace('E+','').isdigit()
            else str(v)
        ).str.strip()
        price = pd.to_numeric(df.iloc[:, price_col], errors='coerce')
        tmp = pd.DataFrame({'UPC': upc, 'Price': price, 'Date': date})
        tmp = tmp.dropna(subset=['Price'])
        tmp = tmp[tmp['UPC'].str.len() >= 8]
        rows.append(tmp)
    if not rows:
        return {}
    combined = pd.concat(rows, ignore_index=True).sort_values('Date', ascending=False)
    deduped  = combined.drop_duplicates(subset='UPC', keep='first')
    return dict(zip(deduped['UPC'], deduped['Price']))


# Raw description records for fallback desc matching: {wholesaler_name: [(desc, upc, price)]}
_WS_RAW = {}


def load_wholesaler(cfg):
    """Dispatch to appropriate loader"""
    if cfg['name'] == 'MTZ':
        return load_mtz(cfg)
    elif cfg['name'] == 'Nandansons':
        return load_nandansons(cfg)
    elif cfg['name'] == 'PCA':
        return load_pca(cfg)
    elif cfg['name'] == 'GE':
        return load_ge(cfg)
    elif cfg['name'] == 'PTC':
        return load_ptc(cfg)
    else:
        return load_generic(cfg)


def load_generic(cfg):
    """Generic loader for any new wholesaler with simple structure"""
    lookup = {}
    for f in cfg['files']:
        if not os.path.exists(f): continue
        sheet = cfg.get('sheet', None)
        df = pd.read_excel(f, sheet_name=sheet, header=cfg['header_row'])
        df = df.dropna(subset=[df.columns[cfg['upc_col']]])
        upc_col   = df.columns[cfg['upc_col']]
        price_col = df.columns[cfg['price_col']]
        df[upc_col]   = df[upc_col].astype(str).str.strip()
        df[price_col] = pd.to_numeric(df[price_col], errors='coerce')
        for _, row in df.dropna(subset=[price_col]).iterrows():
            upc = row[upc_col]
            if upc not in lookup:
                lookup[upc] = float(row[price_col])
    return lookup


def detect_col(df, keywords):
    """Find column index by keyword match in column names"""
    for i, col in enumerate(df.columns):
        if any(k.lower() in str(col).lower() for k in keywords):
            return i
    return None


def load_offer(filepath):
    """Load the offer/inventory file. Returns DataFrame with UPC, Price, Description, QTY"""
    df = pd.read_excel(filepath, header=None)
    # Find header row (row containing 'UPC' or 'EAN' or 'Item')
    header_row = 0
    for i, row in df.iterrows():
        vals = [str(v).lower() for v in row.values if pd.notna(v)]
        if any('upc' in v or 'ean' in v or 'item' in v for v in vals):
            header_row = i
            break
    df = pd.read_excel(filepath, header=header_row)
    df.columns = [str(c).strip() for c in df.columns]

    upc_i   = detect_col(df, ['upc','ean','barcode','item number','item_number','ean code'])
    price_i = detect_col(df, ['price','unit price','cost','prix','price usd'])
    desc_i  = detect_col(df, ['description','desc','name','product','designation','item name','title'])
    qty_i   = detect_col(df, ['qty','quantity','qtty','stock','units'])

    result = pd.DataFrame()
    if upc_i is not None:
        # Handle numeric UPCs that get loaded as floats (e.g. 3386460078337.0 → 3386460078337)
        raw = df.iloc[:,upc_i]
        result['UPC'] = raw.apply(
            lambda v: str(int(float(v))) if pd.notna(v) and str(v).replace('.','').replace('e+','').isdigit() or (isinstance(v, float) and not pd.isna(v))
            else str(v)
        ).str.strip()
    if price_i is not None:
        result['My_Price'] = pd.to_numeric(df.iloc[:,price_i], errors='coerce')
    if desc_i is not None:
        result['Description'] = df.iloc[:,desc_i].astype(str)
    if qty_i is not None:
        result['QTY'] = pd.to_numeric(df.iloc[:,qty_i], errors='coerce')

    # Remove header-like rows that snuck in
    if 'UPC' in result.columns:
        result = result[~result['UPC'].str.lower().isin(['upc','ean','item number','nan'])]
    result = result.dropna(subset=['UPC'] if 'UPC' in result.columns else result.columns[:1])
    return result.reset_index(drop=True)



# ════════════════════════════════════════════════════════════════
# DESCRIPTION-BASED MATCHING — fallback when no EAN match
# ════════════════════════════════════════════════════════════════

# Words to ignore when building keyword search tokens
_SKIP = {
    'spray','tester','ladies','eau','edt','edp','parfum','for','the','and',
    'with','women','men','woman','man','de','du','le','la','les','pour',
    'body','lotion','gel','shower','gift','100ml','50ml','75ml',
    '30ml','90ml','125ml','150ml','200ml','60ml','80ml','40ml','10ml',
    '1oz','2oz','3oz','3.4oz','1.7oz','refill','rfl','vap','vapo',
    'intense','intensely','parfum','cologne','toilette','new','nf','np',
    'ml','oz','sp','lc','fg','ad','mad','rev','reno','mv','tr','b',
}
# Extra tokens that indicate a SET/BUNDLE — used to reject these from single-item matching
_SET_TOKENS = {'set','piece','bundle','kit','trio','duo','pack','collection','travel'}

def _tokens(s):
    """Extract meaningful search tokens from a description string"""
    import re
    words = re.sub(r'[^a-z0-9 ]', ' ', str(s).lower()).split()
    return [w for w in words if w not in _SKIP and len(w) >= 3]


def _extract_ml(s):
    """Extract the primary ml/oz size from a description string. Returns float ml or None."""
    import re
    s = str(s).lower()
    # Match patterns like 100ml, 100 ml, 3.4oz, 3.4 oz
    ml_match = re.search(r'(\d+(?:\.\d+)?)\s*ml', s)
    oz_match  = re.search(r'(\d+(?:\.\d+)?)\s*oz', s)
    if ml_match:
        return float(ml_match.group(1))
    if oz_match:
        return round(float(oz_match.group(1)) * 29.5735, 0)  # oz → ml
    return None


def _size_ok(query_ml, candidate_ml, tolerance=0.15):
    """Return True if sizes are compatible (within tolerance, or either is None)."""
    if query_ml is None or candidate_ml is None:
        return True  # no size info — don't filter
    return abs(query_ml - candidate_ml) / max(query_ml, candidate_ml) <= tolerance


def build_desc_index(wholesaler_lookups, wholesaler_names, wholesaler_raw):
    """
    Build a searchable index: token_set → {wholesaler: price}
    wholesaler_raw: list of (name, [(desc, upc, price), ...])
    Returns: list of (token_frozenset, wholesaler_name, price, desc, upc)
    """
    index = []
    for wname, records in wholesaler_raw:
        for desc, upc, price in records:
            toks = frozenset(_tokens(desc))
            if len(toks) >= 2:
                index.append((toks, wname, price, desc, upc))
    return index


def desc_match(query_desc, index, min_overlap=2):
    """
    Find best price matches for a description across all wholesalers.
    Returns dict: {wholesaler_name: best_price} for matches with enough overlap.
    """
    q_toks = set(_tokens(query_desc))
    if len(q_toks) < 2:
        return {}

    # Score each index entry
    scored = []
    for toks, wname, price, desc, upc in index:
        overlap = len(q_toks & toks)
        if overlap >= min_overlap:
            scored.append((overlap, len(toks), wname, price, desc, upc))

    if not scored:
        return {}

    # For each wholesaler, pick the entry with highest overlap (ties: smallest token set = most specific)
    best = {}
    for overlap, ntoks, wname, price, desc, upc in sorted(scored, key=lambda x: (-x[0], x[1])):
        if wname not in best:
            best[wname] = (price, desc, upc, overlap)

    return {wname: v[0] for wname, v in best.items()},            {wname: v[1] for wname, v in best.items()},            {wname: v[2] for wname, v in best.items()}


# ════════════════════════════════════════════════════════════════
# EXCEL OUTPUT STYLES
# ════════════════════════════════════════════════════════════════
HDR_FILL  = PatternFill('solid', start_color='1A3A5C')
ALT_FILL  = PatternFill('solid', start_color='F4F7FB')
GREEN     = PatternFill('solid', start_color='C6EFCE')
RED       = PatternFill('solid', start_color='FFC7CE')
YELLOW    = PatternFill('solid', start_color='FFEB9C')
GRAY      = PatternFill('solid', start_color='EFEFEF')
FOUND_CLR = PatternFill('solid', start_color='C6EFCE')
PART_CLR  = PatternFill('solid', start_color='FFEB9C')
MISS_CLR  = PatternFill('solid', start_color='EFEFEF')

# Wholesaler header colors (cycling palette)
WS_COLORS = ['2E7D32','1565C0','6A1B9A','BF360C','00695C','4E342E','283593','37474F']

HDR_FONT  = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
BOLD      = Font(name='Calibri', bold=True, size=10)
REG       = Font(name='Calibri', size=10)
SM        = Font(name='Calibri', size=9)
CENTER    = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT      = Alignment(horizontal='left',   vertical='center')
RIGHT     = Alignment(horizontal='right',  vertical='center')
thin      = Side(style='thin',   color='CCCCCC')
med       = Side(style='medium', color='999999')
BDR       = Border(left=thin, right=thin, top=thin, bottom=thin)
MED_BDR   = Border(left=med,  right=med,  top=med,  bottom=med)


def pct_font(val):
    if not isinstance(val, (int, float)): return SM
    color = '375623' if val < 0 else ('C00000' if val > 0 else '000000')
    return Font(name='Calibri', bold=True, size=9, color=color)


def build_report(offer_df, wholesaler_lookups, wholesaler_names, output_path):
    """Build the master Excel price analysis report"""
    n_ws = len(wholesaler_names)

    # Merge prices — preserve any desc-matched prices already set
    for wname, lookup in zip(wholesaler_names, wholesaler_lookups):
        col = f'price_{wname}'
        if col not in offer_df.columns:
            offer_df[col] = offer_df['UPC'].map(lookup)
        else:
            # Only fill NaN slots from UPC lookup (keep desc-matched values)
            upc_prices = offer_df['UPC'].map(lookup)
            offer_df[col] = offer_df[col].combine_first(upc_prices)
        offer_df[f'diff_{wname}']  = (offer_df['My_Price'] - offer_df[col]).round(2)
        offer_df[f'pct_{wname}']   = ((offer_df['My_Price'] - offer_df[col]) 
                                       / offer_df[col] * 100).round(1)

    # Industry average
    price_cols = [f'price_{w}' for w in wholesaler_names]
    offer_df['Avg_Price'] = offer_df[price_cols].mean(axis=1, skipna=True).round(2)
    offer_df['Avg_Diff']  = (offer_df['My_Price'] - offer_df['Avg_Price']).round(2)
    offer_df['Avg_Pct']   = ((offer_df['My_Price'] - offer_df['Avg_Price']) 
                              / offer_df['Avg_Price'] * 100).round(1)
    offer_df['Suppliers_Found'] = offer_df[price_cols].notna().sum(axis=1)

    wb = Workbook()

    # ── SHEET 1: FULL ANALYSIS ───────────────────────────────────
    ws = wb.active
    ws.title = 'Full Analysis'
    _write_analysis_sheet(ws, offer_df, wholesaler_names, title=f'Master Price Analysis — {os.path.basename(output_path).replace(".xlsx","")}')

    # ── SHEET 2: MATCHED ONLY ────────────────────────────────────
    ws2 = wb.create_sheet('Matched Items')
    matched = offer_df[offer_df['Suppliers_Found'] > 0].copy()
    _write_analysis_sheet(ws2, matched, wholesaler_names,
                          title=f'Matched Items — {len(matched)} of {len(offer_df)} items found in at least 1 wholesaler')

    # ── SHEET 3: NOT FOUND ───────────────────────────────────────
    ws3 = wb.create_sheet('Not Found')
    notfound = offer_df[offer_df['Suppliers_Found'] == 0].copy()
    _write_notfound_sheet(ws3, notfound)

    # ── SHEET 4: SUMMARY DASHBOARD ───────────────────────────────
    ws4 = wb.create_sheet('Summary', 0)  # Insert at beginning
    _write_summary_sheet(ws4, offer_df, wholesaler_names)

    # Re-order: Summary first
    wb.move_sheet('Summary', offset=-wb.index(wb['Summary']))

    wb.save(output_path)
    print(f"✅ Report saved: {output_path}")
    print(f"   {len(offer_df)} items | {len(matched)} matched | {len(notfound)} not found")
    for wname, lookup in zip(wholesaler_names, wholesaler_lookups):
        cnt = offer_df[f'price_{wname}'].notna().sum()
        print(f"   {wname}: {cnt} matches")


def _write_analysis_sheet(ws, df, wholesaler_names, title):
    n_ws = len(wholesaler_names)

    # ── Build column schema ──────────────────────────────────────
    # Fixed cols: UPC, Description, QTY, My Price, Status
    # Per wholesaler: Price, Diff$, Diff%
    # Final: Avg Price, Diff$, Diff%, Suppliers#
    fixed_cols = ['UPC', 'Description', 'QTY', 'My Price ($)', 'Match Status']
    ws_col_groups = []
    for i, wname in enumerate(wholesaler_names):
        ws_col_groups.append((wname, [f'{wname} Price ($)', f'Diff ($)', f'Diff (%)']))
    avg_cols = ['Avg Price ($)', 'Diff ($)', 'Diff (%)', '# Sources']

    all_headers = fixed_cols.copy()
    for _, cols in ws_col_groups:
        all_headers.extend(cols)
    all_headers.extend(avg_cols)

    total_cols = len(all_headers)

    # ── Title ────────────────────────────────────────────────────
    ws.merge_cells(f'A1:{get_column_letter(total_cols)}1')
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(name='Calibri', bold=True, size=14, color='1A3A5C')
    c.alignment = CENTER
    ws.row_dimensions[1].height = 30

    # ── Sub info ─────────────────────────────────────────────────
    ws.merge_cells(f'A2:{get_column_letter(total_cols)}2')
    c = ws.cell(row=2, column=1,
        value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  "
              f"Items: {len(df)}  |  Wholesalers: {', '.join(wholesaler_names)}  |  "
              f"Green % = your price is BELOW wholesaler (you buy cheaper)  |  "
              f"Red % = your price is ABOVE wholesaler (you pay more)")
    c.font = Font(name='Calibri', size=9, italic=True, color='666666')
    c.alignment = LEFT
    ws.row_dimensions[2].height = 14

    # ── Group headers row 3 ───────────────────────────────────────
    col_cursor = 1
    # Fixed group
    ws.merge_cells(f'A3:{get_column_letter(len(fixed_cols))}3')
    c = ws.cell(row=3, column=1, value='OFFER / INVENTORY')
    c.font = Font(name='Calibri', bold=True, size=9, color='FFFFFF')
    c.fill = PatternFill('solid', start_color='1A3A5C')
    c.alignment = CENTER; c.border = BDR
    col_cursor += len(fixed_cols)

    # Wholesaler groups
    for i, (wname, wcols) in enumerate(ws_col_groups):
        sc = col_cursor; ec = col_cursor + len(wcols) - 1
        ws.merge_cells(f'{get_column_letter(sc)}3:{get_column_letter(ec)}3')
        c = ws.cell(row=3, column=sc, value=wname)
        c.font = Font(name='Calibri', bold=True, size=9, color='FFFFFF')
        c.fill = PatternFill('solid', start_color=WS_COLORS[i % len(WS_COLORS)])
        c.alignment = CENTER; c.border = BDR
        col_cursor += len(wcols)

    # Avg group
    sc = col_cursor; ec = col_cursor + len(avg_cols) - 1
    ws.merge_cells(f'{get_column_letter(sc)}3:{get_column_letter(ec)}3')
    c = ws.cell(row=3, column=sc, value='INDUSTRY AVERAGE')
    c.font = Font(name='Calibri', bold=True, size=9, color='FFFFFF')
    c.fill = PatternFill('solid', start_color='4A0072')
    c.alignment = CENTER; c.border = BDR
    ws.row_dimensions[3].height = 16

    # ── Column headers row 4 ─────────────────────────────────────
    col_widths = [18, 42, 7, 13, 13]
    for _ in wholesaler_names:
        col_widths.extend([13, 12, 10])
    col_widths.extend([13, 12, 10, 10])

    for col, (h, w) in enumerate(zip(all_headers, col_widths), 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = HDR_FONT; c.fill = HDR_FILL
        c.alignment = CENTER; c.border = BDR
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[4].height = 30

    # ── Data rows ────────────────────────────────────────────────
    for ri, (_, row) in enumerate(df.iterrows(), 5):
        alt = ri % 2 == 0
        found = int(row['Suppliers_Found']) if pd.notna(row['Suppliers_Found']) else 0
        n_total = len(wholesaler_names)
        if found == n_total:   status = f'ALL {n_total}'
        elif found > 0:        status = f'{found}/{n_total} FOUND'
        else:                  status = 'NOT FOUND'

        fp = lambda v: v if pd.notna(v) else '—'

        row_vals = [
            row.get('UPC',''),
            row.get('Description',''),
            row.get('QTY',''),
            fp(row.get('My_Price')),
            status,
        ]
        for wname in wholesaler_names:
            row_vals.extend([
                fp(row[f'price_{wname}']),
                fp(row[f'diff_{wname}']),
                fp(row[f'pct_{wname}']),
            ])
        row_vals.extend([
            fp(row['Avg_Price']),
            fp(row['Avg_Diff']),
            fp(row['Avg_Pct']),
            int(found) if found else '—',
        ])

        for col, val in enumerate(row_vals, 1):
            c = ws.cell(row=ri, column=col, value=val)
            c.font = REG; c.border = BDR

            header = all_headers[col-1]
            is_alt = alt

            if 'Price' in header and header != 'Match Status':
                if isinstance(val, float): c.number_format = '$#,##0.00'
                c.alignment = RIGHT
                if isinstance(val, str) and val == '—': c.fill = GRAY
                elif is_alt: c.fill = ALT_FILL

            elif 'Diff ($)' in header:
                if isinstance(val, float):
                    c.number_format = '$#,##0.00'
                    c.fill = GREEN if val < 0 else (RED if val > 0 else PatternFill())
                else:
                    c.fill = GRAY
                c.alignment = RIGHT

            elif 'Diff (%)' in header:
                if isinstance(val, float):
                    c.number_format = '0.0"%"'
                    c.fill = GREEN if val < 0 else (RED if val > 0 else PatternFill())
                    c.font = pct_font(val)
                else:
                    c.fill = GRAY
                c.alignment = CENTER

            elif header == 'Match Status':
                c.alignment = CENTER
                if 'ALL' in str(val):
                    c.fill = FOUND_CLR
                    c.font = Font(name='Calibri', bold=True, size=9, color='375623')
                elif 'FOUND' in str(val):
                    c.fill = PART_CLR
                    c.font = Font(name='Calibri', size=9, color='9C5700')
                else:
                    c.fill = MISS_CLR
                    c.font = Font(name='Calibri', size=9, color='888888')

            elif header == '# Sources':
                c.alignment = CENTER
                if is_alt: c.fill = ALT_FILL

            elif header == 'QTY':
                c.alignment = RIGHT
                if is_alt: c.fill = ALT_FILL

            else:
                c.alignment = LEFT
                if is_alt: c.fill = ALT_FILL

    ws.freeze_panes = 'A5'
    ws.auto_filter.ref = f'A4:{get_column_letter(total_cols)}{4+len(df)}'


def _write_notfound_sheet(ws, df):
    ws.merge_cells('A1:D1')
    c = ws.cell(row=1, column=1, value=f'Items Not Found in Any Wholesaler ({len(df)} items)')
    c.font = Font(name='Calibri', bold=True, size=13, color='C00000')
    c.alignment = CENTER; ws.row_dimensions[1].height = 24

    hdrs = ['UPC', 'Description', 'QTY', 'My Price ($)']
    widths = [18, 55, 8, 14]
    for col, (h, w) in enumerate(zip(hdrs, widths), 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = HDR_FONT; c.fill = HDR_FILL
        c.alignment = CENTER; c.border = BDR
        ws.column_dimensions[get_column_letter(col)].width = w

    for ri, (_, row) in enumerate(df.iterrows(), 3):
        for col, val in enumerate([
            row.get('UPC',''), row.get('Description',''),
            row.get('QTY',''), row.get('My_Price','')
        ], 1):
            c = ws.cell(row=ri, column=col, value=val)
            c.font = REG; c.border = BDR
            c.alignment = RIGHT if col in (3,4) else LEFT
            if col == 4 and isinstance(val, float): c.number_format = '$#,##0.00'

    ws.freeze_panes = 'A3'
    ws.auto_filter.ref = f'A2:D{2+len(df)}'


def _write_summary_sheet(ws, df, wholesaler_names):
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18

    ws.merge_cells('A1:D1')
    c = ws.cell(row=1, column=1, value='📊 PRICE ANALYSIS SUMMARY')
    c.font = Font(name='Calibri', bold=True, size=16, color='1A3A5C')
    c.alignment = CENTER; ws.row_dimensions[1].height = 36

    ws.merge_cells('A2:D2')
    c = ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%B %d, %Y %H:%M')}")
    c.font = Font(name='Calibri', size=10, italic=True, color='888888')
    c.alignment = CENTER; ws.row_dimensions[2].height = 18

    r = 4
    def section(title):
        nonlocal r
        ws.merge_cells(f'A{r}:D{r}')
        c = ws.cell(row=r, column=1, value=title)
        c.font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
        c.fill = PatternFill('solid', start_color='1A3A5C')
        c.alignment = LEFT
        r += 1

    def stat_row(label, *values):
        nonlocal r
        ws.cell(row=r, column=1, value=label).font = REG
        for i, val in enumerate(values, 2):
            c = ws.cell(row=r, column=i, value=val)
            c.font = BOLD
            if isinstance(val, float) and '$' in label: c.number_format = '$#,##0.00'
        r += 1

    section('OVERVIEW')
    total = len(df)
    matched = int(df['Suppliers_Found'].gt(0).sum())
    notfound = total - matched
    stat_row('Total items in offer', total)
    stat_row('Items matched (≥1 wholesaler)', matched, f'{matched/total*100:.1f}%' if total else '—')
    stat_row('Items not found in any wholesaler', notfound, f'{notfound/total*100:.1f}%' if total else '—')
    r += 1

    section('BY WHOLESALER')
    ws.cell(row=r, column=1, value='Wholesaler').font = Font(name='Calibri', bold=True, size=10)
    ws.cell(row=r, column=2, value='Matches').font = Font(name='Calibri', bold=True, size=10)
    ws.cell(row=r, column=3, value='Cheaper Than').font = Font(name='Calibri', bold=True, size=10)
    ws.cell(row=r, column=4, value='More Expensive').font = Font(name='Calibri', bold=True, size=10)
    r += 1
    for i, wname in enumerate(wholesaler_names):
        cnt  = int(df[f'price_{wname}'].notna().sum())
        cheaper = int((df[f'diff_{wname}'] < 0).sum())
        pricier = int((df[f'diff_{wname}'] > 0).sum())
        ws.cell(row=r, column=1, value=wname).font = Font(name='Calibri', bold=True, size=10,
            color=WS_COLORS[i % len(WS_COLORS)])
        ws.cell(row=r, column=2, value=cnt).font = BOLD
        c3 = ws.cell(row=r, column=3, value=cheaper)
        c3.font = Font(name='Calibri', bold=True, color='375623')
        c4 = ws.cell(row=r, column=4, value=pricier)
        c4.font = Font(name='Calibri', bold=True, color='C00000')
        r += 1
    r += 1

    section('VS INDUSTRY AVERAGE (matched items only)')
    matched_df = df[df['Suppliers_Found'] > 0]
    if len(matched_df):
        above = int((matched_df['Avg_Diff'] > 0).sum())
        below = int((matched_df['Avg_Diff'] < 0).sum())
        equal = int((matched_df['Avg_Diff'] == 0).sum())
        avg_pct = matched_df['Avg_Pct'].mean()
        stat_row('Your price ABOVE industry avg', above)
        stat_row('Your price BELOW industry avg', below)
        stat_row('Your price equal to avg', equal)
        stat_row(f'Avg % vs industry avg', f'{avg_pct:+.1f}%')
    r += 1

    section('HOW TO USE')
    tips = [
        ('Green % cells', 'Your price is LOWER than that wholesaler — good deal / competitive'),
        ('Red % cells',   'Your price is HIGHER than that wholesaler — may lose margin'),
        ('— (dash)',       'UPC not found in that wholesaler\'s pricelist'),
        ('# Sources col', 'How many wholesalers have this item listed'),
        ('Ind. Avg',      'Mean of all wholesaler prices where found'),
    ]
    for label, tip in tips:
        c1 = ws.cell(row=r, column=1, value=label); c1.font = BOLD
        ws.merge_cells(f'B{r}:D{r}')
        c2 = ws.cell(row=r, column=2, value=tip); c2.font = REG
        r += 1


# ════════════════════════════════════════════════════════════════
# PACKAGE PRICING — equal-competitiveness redistribution
# ════════════════════════════════════════════════════════════════

TARGET_DISCOUNT = 30.0  # % below industry avg needed (15% tariff + 5% ship + 10-15% margin)

def build_package_price_report(offer_df, wholesaler_names, analysis_path):
    """
    Redistribute prices so every matched item sits at the same % discount
    vs its industry average, preserving total package value.
    Always shows Target Price column (Avg * 0.70) to hit −30%.
    Layout matches reference file exactly (11 columns, single sheet).
    """
    matched   = offer_df[offer_df['Suppliers_Found'] > 0].copy()
    unmatched = offer_df[offer_df['Suppliers_Found'] == 0].copy()

    if len(matched) == 0:
        print("   ⚠️  No matched items — cannot build package price.")
        return None

    if 'QTY' not in matched.columns or matched['QTY'].isna().all():
        matched['QTY'] = 1
    matched['QTY'] = matched['QTY'].fillna(1)

    total_value_matched = (matched['My_Price'] * matched['QTY']).sum()
    total_avg_value     = (matched['Avg_Price'] * matched['QTY']).sum()

    if total_avg_value == 0:
        print("   ⚠️  Industry avg total is zero — cannot compute k.")
        return None

    # Three target tiers: 30% minimum, 35% good, 40% sharp
    TIERS = [
        (30.0, 'Minimum', 'E65100'),
        (35.0, 'Good',    '1565C0'),
        (40.0, 'Sharp',   '1B5E20'),
    ]
    k             = total_value_matched / total_avg_value
    discount_pct  = (k - 1) * 100
    meets_target  = discount_pct <= -TARGET_DISCOUNT
    min_pkg_value = total_avg_value * (1.0 - TARGET_DISCOUNT / 100.0)
    gap_to_target = total_value_matched - min_pkg_value

    matched['New_Price']    = (matched['Avg_Price'] * k).round(2)
    matched['New_Pct']      = ((matched['New_Price'] / matched['Avg_Price'] - 1) * 100).round(1)
    matched['Old_Price']    = matched['My_Price']
    matched['Price_Change'] = (matched['New_Price'] - matched['Old_Price']).round(2)
    for pct, label, _ in TIERS:
        matched[f'Target_{pct:.0f}'] = (matched['Avg_Price'] * (1 - pct/100)).round(2)
    matched['Target_Price'] = matched['Target_30']

    new_total     = (matched['New_Price'] * matched['QTY']).sum()
    t30_total     = (matched['Target_30'] * matched['QTY']).sum()
    t35_total     = (matched['Target_35'] * matched['QTY']).sum()
    t40_total     = (matched['Target_40'] * matched['QTY']).sum()
    rounding_diff = total_value_matched - new_total

    # ── STYLES — exactly matching reference file ──────────────────
    F_DARK  = PatternFill('solid', start_color='1A3A5C')
    F_DGRN  = PatternFill('solid', start_color='1B5E20')
    F_C00   = PatternFill('solid', start_color='C00000')
    F_GRN   = PatternFill('solid', start_color='C6EFCE')
    F_RED   = PatternFill('solid', start_color='FFC7CE')
    F_PURP  = PatternFill('solid', start_color='EDE0F5')
    F_TEAL  = PatternFill('solid', start_color='E0F2F1')
    F_ALT   = PatternFill('solid', start_color='F4F7FB')
    F_LGRY  = PatternFill('solid', start_color='F8F8F8')
    F_AMBR  = PatternFill('solid', start_color='FFF3E0')
    F_ORG   = PatternFill('solid', start_color='E65100')
    F_SLATE = PatternFill('solid', start_color='546E7A')

    def fw(sz=10, bold=True,  color='FFFFFF'): return Font(name='Calibri', bold=bold, size=sz, color=color)
    def fb(sz=10, color='000000'): return Font(name='Calibri', bold=True,  size=sz, color=color)
    def fr(sz=10, color='000000'): return Font(name='Calibri', bold=False, size=sz, color=color)
    def fs(sz=9,  color='666666'): return Font(name='Calibri', bold=False, size=sz, color=color, italic=True)

    _t   = Side(style='thin',   color='CCCCCC')
    _m   = Side(style='medium', color='888888')
    BDR  = Border(left=_t, right=_t, top=_t, bottom=_t)
    MBDR = Border(left=_m, right=_m, top=_m, bottom=_m)
    CTR  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    LFT  = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    RGT  = Alignment(horizontal='right',  vertical='center')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Package Pricing'

    def cel(row, col, val, font=None, fill=None, align=None, border=None, fmt=None):
        c = ws.cell(row=row, column=col, value=val)
        if font:   c.font   = font
        if fill:   c.fill   = fill
        if align:  c.alignment = align
        if border: c.border = border
        if fmt:    c.number_format = fmt
        return c

    # 13 columns: A=18,B=44,C=8,D=12,E=12,F=13,G=13,H=11,I=7,J=22,K=13,L=13,M=13
    for i, w in enumerate([18, 44, 8, 12, 12, 13, 13, 11, 7, 22, 13, 13, 13], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # R1 — Title
    ws.merge_cells('A1:M1')
    ws.cell(1, 1, 'PACKAGE PRICING — Equal Competitiveness Model').font = Font(
        name='Calibri', bold=True, size=14, color='1A3A5C')
    ws.cell(1, 1).alignment = CTR
    ws.row_dimensions[1].height = 30

    # R2 — Subtitle
    ws.merge_cells('A2:M2')
    ws.cell(2, 1,
        'Prices redistributed so every item sits at the same % discount vs industry average, '
        'keeping total package value identical.  '
        f'Target: ≥{TARGET_DISCOUNT:.0f}% below market '
        '(covers 15% US tariff + 5% shipping + 10-15% client margin).'
    ).font = fs(9, '666666')
    ws.cell(2, 1).alignment = LFT
    ws.row_dimensions[2].height = 16

    # R4 — Status banner (red if below target, green if meets)
    ws.merge_cells('A4:M4')
    if meets_target:
        banner = (f'✅  PACKAGE MEETS TARGET  |  Discount: {discount_pct:.1f}%  '
                  f'(target: ≥−{TARGET_DISCOUNT:.0f}%)')
        b_fill = F_DGRN
    else:
        banner = (f'⚠️  PACKAGE BELOW TARGET  |  Current discount: {discount_pct:.1f}%  |  '
                  f'Target: ≥−{TARGET_DISCOUNT:.0f}%  |  '
                  f'Need to reduce package value by ${gap_to_target:,.0f}  →  '
                  f'Min package value: ${min_pkg_value:,.0f}')
        b_fill = F_C00
    ws.cell(4, 1, banner).font      = fw(11, True, 'FFFFFF')
    ws.cell(4, 1).fill              = b_fill
    ws.cell(4, 1).alignment         = CTR
    ws.row_dimensions[4].height     = 22

    # R5 — Section header
    ws.merge_cells('A5:M5')
    ws.cell(5, 1, 'PACKAGE SUMMARY').font  = fw(10, True, 'FFFFFF')
    ws.cell(5, 1).fill                     = F_DARK
    ws.cell(5, 1).alignment                = LFT
    ws.row_dimensions[5].height            = 18

    # R6–R10 — Stats (two-column layout, cols 1-2 left, cols 4-5 right)
    stats = [
        ('Matched items repriced',           len(matched),             '',
         'Uniform discount vs market',        f'{discount_pct:+.2f}%', ''),
        ('Unmatched items (price unchanged)', len(unmatched),           '',
         f'Target discount (≥−{TARGET_DISCOUNT:.0f}%)', f'−{TARGET_DISCOUNT:.0f}%', ''),
        ('Original total value (matched)',    total_value_matched,      '$#,##0.00',
         'Target: meets threshold?',          '✅ YES' if meets_target else '❌ NO', ''),
        ('New total value (matched)',         round(new_total, 2),      '$#,##0.00',
         'Min package value for 30% target',  round(min_pkg_value, 2), '$#,##0.00'),
        ('Rounding difference',              round(rounding_diff, 2),  '$#,##0.00',
         'Gap to target (reduce by)',         round(gap_to_target, 2), '$#,##0.00'),
    ]
    for i, (ll, lv, lf, rl, rv, rf) in enumerate(stats):
        r = 6 + i
        cel(r, 1, ll, fr(), None, LFT)
        c2 = cel(r, 2, lv, fb(), None, RGT);  (c2.__setattr__('number_format', lf) if lf else None)
        cel(r, 4, rl, fr(), None, LFT)
        c5 = cel(r, 5, rv, fb(), None, RGT);  (c5.__setattr__('number_format', rf) if rf else None)
        ws.row_dimensions[r].height = 16

    # R11 — Why 30% teal note
    ws.merge_cells('A11:M11')
    ws.cell(11, 1,
        'Why 30%?  Tariff to USA: ~15%  |  Shipping from Europe: ~5%  |  '
        'Client margin needed: ~10–15%  →  Total buffer required: ~30–35%'
    ).font = Font(name='Calibri', size=9, italic=True, color='444444')
    ws.cell(11, 1).fill      = F_TEAL
    ws.cell(11, 1).alignment = LFT
    ws.row_dimensions[11].height = 16

    # R13 — Repriced items green header
    ws.merge_cells('A13:M13')
    ws.cell(13, 1, 'REPRICED ITEMS').font  = fw(10, True, 'FFFFFF')
    ws.cell(13, 1).fill                    = F_DGRN
    ws.cell(13, 1).alignment               = LFT
    ws.row_dimensions[13].height           = 18

    # R14 — Column headers (cols K/L/M are 3 tier targets)
    tier_fills = {11: PatternFill('solid', start_color='E65100'),   # orange 30%
                  12: PatternFill('solid', start_color='1565C0'),   # blue   35%
                  13: PatternFill('solid', start_color='1B5E20')}   # green  40%
    hdrs = ['EAN/UPC', 'Description', 'QTY', 'Original $', 'Ind. Avg $',
            'New Price $', 'Change $/pc', 'Disc. vs Avg', '# Src', 'Notes',
            'Target −30%\n(Minimum)', 'Target −35%\n(Good)', 'Target −40%\n(Sharp)']
    for col, h in enumerate(hdrs, 1):
        cel(14, col, h, fw(10, True, 'FFFFFF'),
            tier_fills.get(col, F_DARK), CTR, BDR)
    ws.row_dimensions[14].height = 32

    # R15+ — Data rows
    for i, (_, row) in enumerate(matched.iterrows()):
        r    = 15 + i
        fill = F_ALT if i % 2 == 0 else F_LGRY
        chg  = row['Price_Change']

        cel(r, 1,  row.get('UPC',''),         fr(10), fill,  LFT, BDR)
        cel(r, 2,  row.get('Description',''), fr(10), fill,  LFT, BDR)
        cel(r, 3,  row.get('QTY', 1),         fr(10), fill,  RGT, BDR, '#,##0')
        cel(r, 4,  row['Old_Price'],           fr(10), fill,  RGT, BDR, '$#,##0.00')
        cel(r, 5,  row['Avg_Price'],           fr(10), fill,  RGT, BDR, '$#,##0.00')
        # F: new price — always green fill, bold black
        cel(r, 6,  row['New_Price'],           fb(10), F_GRN, RGT, BDR, '$#,##0.00')
        # G: change — green fill+font if raised, red if lowered
        chg_fill = F_GRN if chg > 0 else (F_RED if chg < 0 else fill)
        chg_font = fb(10, '1B5E20') if chg > 0 else (fb(10, 'C00000') if chg < 0 else fr(10))
        cel(r, 7,  chg, chg_font, chg_fill, RGT, BDR, '+$#,##0.00;-$#,##0.00;$-')
        # H: disc% — green fill, green font (model always gives same negative %)
        cel(r, 8,  row['New_Pct'] / 100, fb(10, '1B5E20'), F_GRN, CTR, BDR, '+0.0%;-0.0%;0.0%')
        # I: sources count
        cel(r, 9,  int(row['Suppliers_Found']), fr(10), fill, CTR, BDR)
        # J: notes — small italic grey
        if abs(chg) < 0.01: note = 'Unchanged'
        elif chg > 0:       note = f'↑ +${chg:.2f}/pc'
        else:               note = f'↓ −${abs(chg):.2f}/pc'
        cel(r, 10, note, fs(9, '666666'), fill, LFT, BDR)
        # K/L/M: three target tier prices
        tier_data = [
            (11, 'Target_30', 'E65100', PatternFill('solid', start_color='FFF3E0')),
            (12, 'Target_35', '1565C0', PatternFill('solid', start_color='E3F2FD')),
            (13, 'Target_40', '1B5E20', PatternFill('solid', start_color='E8F5E9')),
        ]
        for tcol, tkey, tcolor, tfill in tier_data:
            cel(r, tcol, row[tkey], fb(10, tcolor), tfill, RGT, BDR, '$#,##0.00')
        ws.row_dimensions[r].height = 18

    # Totals row
    r_tot = 15 + len(matched)
    cel(r_tot, 1,  'TOTAL',                               fb(10), F_PURP, LFT,  MBDR)
    cel(r_tot, 2,  '',                                    fr(10), F_PURP, LFT,  MBDR)
    cel(r_tot, 3,  int(matched['QTY'].sum()),              fb(10), F_PURP, RGT,  MBDR, '#,##0')
    cel(r_tot, 4,  round(total_value_matched, 2),          fb(10), F_PURP, RGT,  MBDR, '$#,##0.00')
    cel(r_tot, 5,  round(total_avg_value, 2),              fb(10), F_PURP, RGT,  MBDR, '$#,##0.00')
    cel(r_tot, 6,  round(new_total, 2),                    fb(10), F_PURP, RGT,  MBDR, '$#,##0.00')
    cel(r_tot, 7,  round(new_total - total_value_matched, 2), fb(10), F_PURP, RGT, MBDR, '+$#,##0.00;-$#,##0.00;$-')
    # H total: green/red depending on target
    cel(r_tot, 8,  f'{discount_pct:+.2f}%',
        fb(10), F_GRN if meets_target else F_RED, CTR, MBDR)
    cel(r_tot, 9,  '', fr(10), F_PURP, CTR, MBDR)
    cel(r_tot, 10, '', fr(10), F_PURP, LFT, MBDR)
    # K/L/M totals for 3 tiers
    for tcol, ttot, tcolor in [(11, round(t30_total,2), 'E65100'),
                                (12, round(t35_total,2), '1565C0'),
                                (13, round(t40_total,2), '1B5E20')]:
        cel(r_tot, tcol, ttot, fw(10, True, 'FFFFFF'),
            PatternFill('solid', start_color=tcolor), RGT, MBDR, '$#,##0.00')
    ws.row_dimensions[r_tot].height = 22

    # R(tot+2) — callout row
    r_call = r_tot + 2
    ws.merge_cells(f'A{r_call}:M{r_call}')
    if meets_target:
        call_txt  = (f'✅  PACKAGE MEETS −{TARGET_DISCOUNT:.0f}% TARGET  |  '
                     f'Discount: {discount_pct:.1f}%  |  Total value: ${new_total:,.0f}')
        call_fill = F_DGRN
    else:
        call_txt  = (f'📌  TO REACH −{TARGET_DISCOUNT:.0f}% TARGET:  '
                     f'Package value must be ≤ ${min_pkg_value:,.0f}  '
                     f'(current: ${new_total:,.0f}  |  reduce by ${gap_to_target:,.0f})  '
                     f'—  Target prices shown in orange column above.')
        call_fill = F_ORG
    ws.cell(r_call, 1, call_txt).font  = fw(10, True, 'FFFFFF')
    ws.cell(r_call, 1).fill            = call_fill
    ws.cell(r_call, 1).alignment       = LFT
    ws.row_dimensions[r_call].height   = 20

    # Unmatched items section
    if len(unmatched):
        r = r_call + 2
        ws.merge_cells(f'A{r}:M{r}')
        ws.cell(r, 1,
            f'UNMATCHED ITEMS — {len(unmatched)} items  '
            '(no industry avg available — prices unchanged)'
        ).font = fw(10, True, 'FFFFFF')
        ws.cell(r, 1).fill      = F_SLATE
        ws.cell(r, 1).alignment = LFT
        ws.row_dimensions[r].height = 18
        r += 1
        for col, h in enumerate(['EAN/UPC', 'Description', 'QTY', 'Price (unchanged)'], 1):
            cel(r, col, h, fw(10, True, 'FFFFFF'), F_DARK, CTR, BDR)
        ws.row_dimensions[r].height = 20
        for j, (_, row_) in enumerate(unmatched.iterrows()):
            r += 1
            fill = F_ALT if j % 2 == 0 else F_LGRY
            cel(r, 1, row_.get('UPC',''),         fr(10), fill, LFT, BDR)
            cel(r, 2, row_.get('Description',''), fr(10), fill, LFT, BDR)
            cel(r, 3, row_.get('QTY', ''),        fr(10), fill, RGT, BDR, '#,##0')
            cel(r, 4, row_.get('My_Price', ''),   fr(10), fill, RGT, BDR, '$#,##0.00')
            ws.row_dimensions[r].height = 18

    ws.freeze_panes = 'A3'

    base = analysis_path.replace('.xlsx', '_PackagePrice.xlsx')
    wb.save(base)
    status = (f"✅ Meets −{TARGET_DISCOUNT:.0f}% target" if meets_target
              else f"⚠️  Below target — need ${gap_to_target:,.0f} reduction (min pkg: ${min_pkg_value:,.0f})")
    print(f"✅ Package price report: {base}")
    print(f"   Discount vs market:  {discount_pct:+.2f}%  |  {status}")
    print(f"   Total value:         ${new_total:,.2f}  (rounding diff: ${rounding_diff:.2f})")
    return base, matched[['UPC','Description','QTY','Old_Price','Avg_Price','New_Price','New_Pct','Target_Price']]


def _build_ws_raw(cfg):
    """
    Collect (description, upc, price) tuples from a wholesaler config,
    used for description-based fallback matching.
    Returns list of (desc_str, upc_str, price_float).
    """
    def unwrap(f):
        """Handle both plain paths and (path, date) tuples"""
        return f[0] if isinstance(f, tuple) else f

    records = []
    name = cfg['name']

    def safe_upc(v):
        s = str(v).strip().replace('.0','')
        try: return str(int(float(s)))
        except: return None

    def safe_price(v):
        try: p = float(v); return p if p > 0 else None
        except: return None

    def get_desc(row, cols):
        parts = [str(row.iloc[i]) for i in cols if i < len(row) and pd.notna(row.iloc[i]) and str(row.iloc[i]).lower() not in ('nan','none','')]
        return ' '.join(parts)

    files = [unwrap(f) for f in cfg.get('files', [])]

    if name == 'MTZ':
        for fpath in files:
            if not os.path.exists(fpath): continue
            try:
                df = pd.read_excel(fpath, sheet_name=cfg.get('sheet'), header=cfg.get('header_row', 4))
                for _, row in df.iterrows():
                    upc = safe_upc(row.iloc[cfg['upc_col']])
                    price = safe_price(row.iloc[cfg['price_col']])
                    if upc and price:
                        records.append((get_desc(row, [0,1,2]), upc, price))
            except: pass

    elif name == 'Nandansons':
        for fpath in files:
            if not os.path.exists(fpath): continue
            try:
                df = pd.read_excel(fpath, header=0)
                ncols = len(df.columns)
                pc = 7 if ncols >= 8 else 5
                for _, row in df.iterrows():
                    upc = safe_upc(row.iloc[0])
                    price = safe_price(row.iloc[pc])
                    if upc and price:
                        records.append((get_desc(row, [1,2,3]), upc, price))
            except: pass

    elif name == 'PCA':
        for fpath in files:
            if not os.path.exists(fpath): continue
            try:
                df = pd.read_excel(fpath, header=0)
                for _, row in df.iterrows():
                    upc = safe_upc(row.iloc[cfg['upc_col']])
                    price = safe_price(row.iloc[cfg['price_col']])
                    if upc and price:
                        records.append((get_desc(row, [0,1,2,3]), upc, price))
            except: pass

    elif name == 'GE':
        for fpath in files:
            if not os.path.exists(fpath): continue
            try:
                df = pd.read_excel(fpath, header=0)
                for _, row in df.iterrows():
                    upc = safe_upc(row.iloc[cfg['upc_col']])
                    price = safe_price(row.iloc[cfg['price_col']])
                    if upc and price:
                        # GE: col2=Brand, col3=Description — combine both for richer matching
                        records.append((get_desc(row, [2, 3]), upc, price))
            except: pass

    elif name == 'PTC':
        for fpath in files:
            if not os.path.exists(fpath): continue
            for hrow in [6, 7, 8, 11, 0]:
                for sname in ['Price List', 'Sheet1', None]:
                    try:
                        df = pd.read_excel(fpath, sheet_name=sname, header=hrow)
                        if len(df) < 3: continue
                        for _, row in df.iterrows():
                            upc = safe_upc(row.iloc[1])
                            price = safe_price(row.iloc[4])
                            if upc and price:
                                records.append((get_desc(row, [2,3]), upc, price))
                        break
                    except: pass
                else:
                    continue
                break

    # Deduplicate by upc, keep first
    seen = set()
    deduped = []
    for r in records:
        if r[1] not in seen:
            seen.add(r[1])
            deduped.append(r)
    return deduped



# ════════════════════════════════════════════════════════════════
# MAIN
# ════════════════════════════════════════════════════════════════
def run(offer_file, output_file=None, package_price=False):
    print(f"\n📂 Loading offer: {offer_file}")
    offer_df = load_offer(offer_file)
    print(f"   Found {len(offer_df)} items | Cols: {list(offer_df.columns)}")

    wholesaler_lookups = []
    wholesaler_names   = []
    for cfg in WHOLESALERS:
        lookup = load_wholesaler(cfg)
        wholesaler_lookups.append(lookup)
        wholesaler_names.append(cfg['name'])
        print(f"   Loaded {cfg['name']}: {len(lookup):,} UPCs")
        # Build raw desc records for this wholesaler (for desc-fallback matching)
        _WS_RAW[cfg['name']] = _build_ws_raw(cfg)

    if output_file is None:
        base = os.path.splitext(os.path.basename(offer_file))[0]
        output_file = f'/mnt/user-data/outputs/PriceAnalysis_{base}_{datetime.now().strftime("%Y%m%d")}.xlsx'

    # ── DESC FALLBACK: fill unmatched items by name ──────────────
    # First pass: map by UPC
    for wname, lookup in zip(wholesaler_names, wholesaler_lookups):
        offer_df[f'price_{wname}'] = offer_df['UPC'].map(lookup)
    price_cols = [f'price_{w}' for w in wholesaler_names]
    offer_df['Suppliers_Found_UPC'] = offer_df[price_cols].notna().sum(axis=1)

    # Second pass: desc-match for items with no UPC hit
    unmatched_mask = offer_df['Suppliers_Found_UPC'] == 0
    n_unmatched = unmatched_mask.sum()
    if n_unmatched > 0 and 'Description' in offer_df.columns:
        print(f"   🔍 Trying description match for {n_unmatched} unmatched items...")
        # Build desc index from all wholesaler raw records
        desc_index = []
        for wname, lookup in zip(wholesaler_names, wholesaler_lookups):
            # Build a simple desc index from the wholesaler's own raw data
            # We re-load just descriptions for the desc index
            raw_records = _WS_RAW.get(wname, [])
            for desc, upc, price in raw_records:
                toks = frozenset(_tokens(desc))
                if len(toks) >= 2:
                    desc_index.append((toks, wname, price, desc, upc))

        filled = 0
        for idx in offer_df[unmatched_mask].index:
            query = str(offer_df.at[idx, 'Description'])
            q_toks = set(_tokens(query))
            if len(q_toks) < 2:
                continue
            # Score each index entry — with size-aware filtering
            query_ml = _extract_ml(query)
            scored = []
            for toks, wname, price, desc, upc in desc_index:
                overlap = len(q_toks & toks)
                if overlap >= 2:
                    cand_ml = _extract_ml(desc)
                    if _size_ok(query_ml, cand_ml):
                        # Reject sets/bundles when matching a single item
                        cand_words = set(re.sub(r'[^a-z0-9 ]',' ',desc.lower()).split())
                        if not (cand_words & _SET_TOKENS):
                            scored.append((overlap, len(toks), wname, price, desc, upc))
            if not scored:
                continue
            # Best match per wholesaler
            best_per_ws = {}
            for overlap, ntoks, wname, price, desc, upc in sorted(scored, key=lambda x: (-x[0], x[1])):
                if wname not in best_per_ws:
                    best_per_ws[wname] = (price, desc, overlap)
            for wname, (price, match_desc, overlap) in best_per_ws.items():
                offer_df.at[idx, f'price_{wname}'] = price
            # Update description to show it was desc-matched
            best_ws = max(best_per_ws, key=lambda w: best_per_ws[w][2])
            best_match_desc = best_per_ws[best_ws][1]
            offer_df.at[idx, 'Match_Note'] = f'~desc: {best_match_desc[:40]}'
            filled += 1
        if filled > 0:
            print(f"   ✅ Desc-matched {filled} additional items")
            for idx2 in offer_df[unmatched_mask].index:
                note = offer_df.at[idx2, 'Match_Note'] if 'Match_Note' in offer_df.columns else ''
                if note:
                    desc2 = str(offer_df.at[idx2, 'Description'])[:35]
                    print(f"      • {desc2:<36} → {note}")

    # Recalculate averages and counts after desc matching
    offer_df['Avg_Price']       = offer_df[price_cols].mean(axis=1, skipna=True).round(2)
    offer_df['Suppliers_Found'] = offer_df[price_cols].notna().sum(axis=1)

    build_report(offer_df, wholesaler_lookups, wholesaler_names, output_file)

    # ── PACKAGE PRICING STEP ─────────────────────────────────────
    if not package_price:
        print("\n💡 Want to make a package price?")
        print("   Run with --package flag, or answer below:")
        answer = input("   Make package price? (yes/no): ").strip().lower()
        package_price = answer in ('yes', 'y')

    if package_price:
        print("\n📦 Building package price...")
        for wname, lookup in zip(wholesaler_names, wholesaler_lookups):
            if f'price_{wname}' not in offer_df.columns:
                offer_df[f'price_{wname}'] = offer_df['UPC'].map(lookup)
        price_cols = [f'price_{w}' for w in wholesaler_names]
        offer_df['Avg_Price']       = offer_df[price_cols].mean(axis=1, skipna=True).round(2)
        offer_df['Suppliers_Found'] = offer_df[price_cols].notna().sum(axis=1)

        result = build_package_price_report(offer_df, wholesaler_names, output_file)
        if result:
            pkg_path, repriced = result
            print("\n📋 REPRICED ITEMS:")
            print(f"{'Description':<45} {'Old $':>8} {'Avg $':>8} {'New $':>8} {'Disc%':>7}")
            print("-" * 80)
            for _, row in repriced.iterrows():
                print(f"{str(row['Description'])[:44]:<45} ${row['Old_Price']:>7.2f} ${row['Avg_Price']:>7.2f} ${row['New_Price']:>7.2f} {row['New_Pct']:>+6.1f}%")
            return output_file, pkg_path

    return output_file


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python master_price_tool.py <offer_file.xlsx> [output.xlsx] [--package]")
        sys.exit(1)
    pkg_flag = '--package' in sys.argv
    args = [a for a in sys.argv[1:] if a != '--package']
    out = run(args[0], args[1] if len(args) > 1 else None, package_price=pkg_flag)
    if isinstance(out, tuple):
        print(f"\n✅ Analysis → {out[0]}")
        print(f"✅ Package  → {out[1]}")
    else:
        print(f"\n✅ Done → {out}")
