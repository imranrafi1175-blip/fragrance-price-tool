"""
Fragrance Wholesale Price Intelligence — Streamlit Web App
Uses master_price_tool.py as the backend for 100% identical output.
"""

import streamlit as st
import pandas as pd
import os, sys, io, tempfile, shutil, subprocess, importlib.util, traceback
from pathlib import Path

st.set_page_config(page_title="Fragrance Price Intelligence", page_icon="🧴", layout="wide")

# Use writable directories - works locally and on Streamlit Cloud
BASE_DIR   = Path(tempfile.gettempdir()) / "fragrance_tool"
UPLOAD_DIR = BASE_DIR / "uploads"
CLAUDE_DIR = BASE_DIR / "claude"
OUTPUT_DIR = BASE_DIR / "outputs"
for d in [UPLOAD_DIR, CLAUDE_DIR, OUTPUT_DIR]:
    d.mkdir(parents=True, exist_ok=True)

st.markdown("""<style>
.main-title{font-size:2rem;font-weight:700;color:#1A3A5C;}
.sub-title{font-size:1rem;color:#64748B;margin-bottom:1.5rem;}
.stat-box{background:#F4F7FB;border-radius:8px;padding:1rem;text-align:center;border:1px solid #E2E8F0;}
.stat-num{font-size:1.8rem;font-weight:700;color:#1A3A5C;}
.stat-label{font-size:0.8rem;color:#64748B;}
.verdict-sharp{background:#E8F5E9;border-left:4px solid #1B5E20;padding:1rem;border-radius:4px;}
.verdict-good{background:#E3F2FD;border-left:4px solid #1565C0;padding:1rem;border-radius:4px;}
.verdict-min{background:#FFF3E0;border-left:4px solid #E65100;padding:1rem;border-radius:4px;}
.verdict-bad{background:#FEE2E2;border-left:4px solid #C00000;padding:1rem;border-radius:4px;}
</style>""", unsafe_allow_html=True)

st.markdown('<div class="main-title">🧴 Fragrance Price Intelligence</div>', unsafe_allow_html=True)
st.markdown('<div class="sub-title">MTZ · Nandansons · PCA · GE · PTC · 13,500+ UPCs · Instant deal scoring</div>', unsafe_allow_html=True)

# Patch master_price_tool to use our writable paths
def patch_tool_paths(tool_text):
    """Replace hardcoded /mnt/user-data paths with our writable ones"""
    tool_text = tool_text.replace("'/mnt/user-data/uploads'", f"'{UPLOAD_DIR}'")
    tool_text = tool_text.replace("'/home/claude'", f"'{CLAUDE_DIR}'")
    tool_text = tool_text.replace('/mnt/user-data/outputs/', f'{OUTPUT_DIR}/')
    # Replace the input() prompt that would block in Streamlit
    tool_text = tool_text.replace(
        "answer = input(\"   Make package price? (yes/no): \").strip().lower()",
        "answer = 'yes' if package_price else 'no'"
    )
    return tool_text

# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.header("📦 Wholesaler Databases")
    st.caption("Upload once per session. Files stay loaded.")

    ws_uploads = {}
    for grp, label, exts in [
        ("MTZ",        "MTZ files (.xlsx)",            ["xlsx"]),
        ("Nandansons", "Nandansons files (.xlsx/.xls)",["xlsx","xls"]),
        ("PCA",        "PCA file (.xls/.xlsx)",        ["xls","xlsx"]),
        ("GE",         "GE files (.xls/.xlsx)",        ["xls","xlsx"]),
        ("PTC",        "PTC files (.xlsx)",            ["xlsx"]),
    ]:
        st.subheader(grp)
        ws_uploads[grp] = st.file_uploader(label, type=exts, accept_multiple_files=True, key=f"ws_{grp}") or []

    if st.button("💾 Save & Convert Files", type="primary", use_container_width=True):
        saved = 0
        with st.spinner("Saving and converting files..."):
            for grp, files in ws_uploads.items():
                for f in files:
                    dest = UPLOAD_DIR / f.name
                    dest.write_bytes(f.getvalue())
                    saved += 1
                    if f.name.endswith(".xls"):
                        try:
                            subprocess.run(
                                ["libreoffice","--headless","--convert-to","xlsx",
                                 str(dest),"--outdir", str(CLAUDE_DIR)],
                                capture_output=True, timeout=60
                            )
                        except FileNotFoundError:
                            # LibreOffice not available — try Python xlrd→openpyxl
                            try:
                                import xlrd
                                from openpyxl import Workbook
                                wb_xls = xlrd.open_workbook(str(dest))
                                wb_new = Workbook(); wb_new.remove(wb_new.active)
                                for sh in wb_xls.sheets():
                                    ws = wb_new.create_sheet(sh.name)
                                    for r in range(sh.nrows):
                                        for c in range(sh.ncols):
                                            ws.cell(r+1, c+1, sh.cell_value(r, c))
                                out = CLAUDE_DIR / f.name.replace(".xls", ".xlsx")
                                wb_new.save(str(out))
                            except Exception as e:
                                st.warning(f"⚠️ Could not convert {f.name}: {e}")
        st.success(f"✅ {saved} files saved!")

    st.markdown("---")
    st.caption("**Files detected:**")
    counts = {
        "MTZ":        len(list(UPLOAD_DIR.glob("MTZpricelist*.xlsx"))),
        "Nandansons": len(list(UPLOAD_DIR.glob("Nandansons_*.xlsx"))) + len(list(CLAUDE_DIR.glob("Nandansons_*.xlsx"))),
        "PCA":        len(list(CLAUDE_DIR.glob("PRICE-LIST*.xlsx"))),
        "GE":         len(list(CLAUDE_DIR.glob("GE_*.xlsx"))) + len(list(CLAUDE_DIR.glob("WHOLESALE_*.xlsx"))) + len(list(CLAUDE_DIR.glob("Ge_*.xlsx"))),
        "PTC":        len(list(UPLOAD_DIR.glob("PTC_*.xlsx"))),
    }
    for name, cnt in counts.items():
        st.caption(f"{'✅' if cnt > 0 else '❌'} {name}: {cnt} file(s)")
    total_files = sum(counts.values())

# ── Main ──────────────────────────────────────────────────────────────────────
c1, c2, c3 = st.columns([3, 1, 1])
with c1:
    offer_file = st.file_uploader("📂 Upload Offer File (.xlsx)", type=["xlsx","xls"], key="offer")
with c2:
    eur_mode = st.checkbox("EUR → USD", value=False)
    eur_rate = st.number_input("Rate", value=1.16, step=0.01, format="%.2f") if eur_mode else 1.0
with c3:
    do_package = st.checkbox("Package Price", value=True)

if offer_file and st.button("🚀 Run Price Analysis", type="primary", use_container_width=True):

    if total_files == 0:
        st.error("❌ No wholesaler files found — upload them in the sidebar first.")
        st.stop()

    # Load and patch master_price_tool.py from the repo directory
    repo_tool = Path(__file__).parent / "master_price_tool.py"
    if not repo_tool.exists():
        st.error(f"❌ master_price_tool.py not found in repo at {repo_tool}")
        st.stop()

    # Patch paths and write to writable location
    tool_text = repo_tool.read_text()
    patched_text = patch_tool_paths(tool_text)
    patched_tool = CLAUDE_DIR / "master_price_tool.py"
    patched_tool.write_text(patched_text)

    # Save offer file
    offer_tmp = CLAUDE_DIR / f"tmp_offer_{offer_file.name}"
    offer_tmp.write_bytes(offer_file.getvalue())

    # EUR conversion
    if eur_mode and eur_rate != 1.0:
        try:
            df_raw = pd.read_excel(str(offer_tmp), header=None)
            hrow = 0
            for i, row in df_raw.iterrows():
                vals = [str(v).lower() for v in row.values if pd.notna(v)]
                if any(k in ' '.join(vals) for k in ['price','upc','ean']): hrow=i; break
            df = pd.read_excel(str(offer_tmp), header=hrow)
            df.columns = [str(c).strip() for c in df.columns]
            for c in df.columns:
                if any(k in str(c).lower() for k in ['price','cost','prix']):
                    df[c] = pd.to_numeric(df[c], errors='coerce') * eur_rate
                    break
            df.to_excel(str(offer_tmp), index=False)
            st.info(f"✅ EUR × {eur_rate} = USD applied")
        except Exception as e:
            st.warning(f"⚠️ EUR conversion skipped: {e}")

    output_path = OUTPUT_DIR / f"PriceAnalysis_{offer_file.name}"
    pkg_path    = OUTPUT_DIR / f"PriceAnalysis_{offer_file.name.replace('.xlsx','')}_PackagePrice.xlsx"

    with st.spinner("🔍 Running price analysis across 5 wholesalers..."):
        try:
            if str(CLAUDE_DIR) not in sys.path:
                sys.path.insert(0, str(CLAUDE_DIR))
            if "master_price_tool" in sys.modules:
                del sys.modules["master_price_tool"]
            spec = importlib.util.spec_from_file_location("master_price_tool", str(patched_tool))
            tool = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(tool)
            tool.run(str(offer_tmp), str(output_path), package_price=do_package)
            success = True
        except Exception as e:
            st.error(f"❌ Error: {e}")
            st.code(traceback.format_exc())
            success = False

    if success:
        st.markdown("---")
        st.subheader("✅ Analysis Complete!")

        # Stats
        try:
            wb_r = pd.read_excel(str(output_path), sheet_name=None)
            sheets = list(wb_r.keys())
            matched_sheet  = [s for s in sheets if 'match' in s.lower() and 'not' not in s.lower()]
            notfound_sheet = [s for s in sheets if 'not' in s.lower() and 'found' in s.lower()]
            n_matched  = len(wb_r[matched_sheet[0]])  - 4 if matched_sheet  else 0
            n_notfound = len(wb_r[notfound_sheet[0]]) - 4 if notfound_sheet else 0
            n_matched  = max(0, n_matched)
            n_notfound = max(0, n_notfound)

            cc1, cc2, cc3 = st.columns(3)
            with cc1: st.markdown(f'<div class="stat-box"><div class="stat-num" style="color:#1B5E20">{n_matched}</div><div class="stat-label">Matched</div></div>', unsafe_allow_html=True)
            with cc2: st.markdown(f'<div class="stat-box"><div class="stat-num" style="color:#C00000">{n_notfound}</div><div class="stat-label">Not Found</div></div>', unsafe_allow_html=True)
            with cc3:
                if pkg_path.exists():
                    try:
                        df_b = pd.read_excel(str(pkg_path), sheet_name='Package Pricing', header=None)
                        banner = str(df_b.iloc[3, 0]) if len(df_b) > 3 else ''
                        if 'Discount:' in banner or 'discount:' in banner:
                            txt = banner.replace('discount','Discount')
                            disc = txt.split('Discount:')[1].split('%')[0].strip()
                            color = '#1B5E20' if '✅' in banner else '#C00000'
                            st.markdown(f'<div class="stat-box"><div class="stat-num" style="color:{color}">{disc}%</div><div class="stat-label">Package Discount</div></div>', unsafe_allow_html=True)
                    except: pass
        except Exception as e:
            st.warning(f"Could not parse stats: {e}")

        # Verdict
        if pkg_path.exists():
            try:
                df_b = pd.read_excel(str(pkg_path), sheet_name='Package Pricing', header=None)
                banner = str(df_b.iloc[3, 0]) if len(df_b) > 3 else ''
                st.markdown("")
                if '✅' in banner and 'Discount:' in banner:
                    disc_str = banner.split('Discount:')[1].split('%')[0].strip()
                    disc_num = float(disc_str.replace('+','').replace('−','-'))
                    if disc_num <= -40:
                        st.markdown(f'<div class="verdict-sharp">🟢 <strong>SHARP — Excellent deal!</strong> {disc_str}% below market. Strong buy.</div>', unsafe_allow_html=True)
                    elif disc_num <= -35:
                        st.markdown(f'<div class="verdict-good">🔵 <strong>GOOD — Solid deal.</strong> {disc_str}% below market. Worth taking.</div>', unsafe_allow_html=True)
                    else:
                        st.markdown(f'<div class="verdict-min">🟠 <strong>MINIMUM — Viable.</strong> {disc_str}% below market. Proceed with caution.</div>', unsafe_allow_html=True)
                elif '⚠️' in banner:
                    st.markdown(f'<div class="verdict-bad">⚠️ <strong>BELOW TARGET</strong> — Does not meet −30% minimum. Counter or pass.</div>', unsafe_allow_html=True)
            except: pass

        # Downloads
        st.markdown("### 📥 Download Reports")
        dl1, dl2 = st.columns(2)
        with dl1:
            if output_path.exists():
                st.download_button(
                    "📋 Price Analysis (.xlsx)",
                    data=output_path.read_bytes(),
                    file_name=f"PriceAnalysis_{offer_file.name}",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
        with dl2:
            if do_package and pkg_path.exists():
                st.download_button(
                    "📦 Package Price (.xlsx)",
                    data=pkg_path.read_bytes(),
                    file_name=f"PackagePrice_{offer_file.name}",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

    if offer_tmp.exists():
        offer_tmp.unlink()

st.markdown("---")
st.caption("🟢 −40%+ Sharp  ·  🔵 −35% Good  ·  🟠 −30% Minimum  ·  ⚠️ Below target  ·  🔴 Above market")
